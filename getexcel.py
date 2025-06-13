from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

wb = load_workbook("专利41-80.xlsx")
sheet = wb["氮化镓相关"]

# head_row = next(sheet.iter_rows(min_row=1,max_row=1,values_only=True))
#
# try:
#     col_index = head_row.index("D")
# except ValueError:
#     raise ValueError(f"不存在")
list0 =[]
for row in sheet['A2':'G42']:
    number = 0
    list = []
    for cell in  row:
        number += 1

        if number == 1:
            list.append(cell.value)
            #print(cell.value)
        elif number == 4:
            list.append(cell.value)
        elif number == 7:
            new_name_1 = cell.value.replace("|",",")
            new_name = new_name_1.replace(" ","")
            list.append(new_name)
            #print(new_name)

        else:
            pass
    #print(list)
    list0.append(list)

print(list0)
wb =Workbook()
ws = wb.active
ws.title = "shuju"

headers = ["序号","姓名","题目"]
for col_idx, header in enumerate(headers, start=1):  # 列索引从1开始
    cell = ws.cell(row=1, column=col_idx, value=header)

data=list0
for row_idx, row_data in enumerate(data, start=2):  # 行索引从2开始（跳过表头）
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# 调整列宽（自动适应内容）
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

# 保存文件
wb.save("openpyxl_output.xlsx")
